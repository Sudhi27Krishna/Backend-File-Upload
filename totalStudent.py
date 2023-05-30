import openpyxl
import sys
import json

data_json = sys.argv[1]
data = json.loads(data_json)

details = data['details']
s = set()
for i in range(len(details)):
    s.add(details[i].get("slot"))
slot_list = list(s)
noOfStudents = 0
for i in slot_list:
    input_slot = i
    p = 1
    b = 1
    code_list = list()
    for i in range(len(details)):
        if details[i].get("slot") == input_slot:
            code_list.append(details[i].get("branch"))

    print(code_list)
    for code in code_list:
        check_supply = 1
        wb_branch = openpyxl.load_workbook('.\\updatedExcels\\'+code+'.xlsx')
        ws_branch_reg = wb_branch[input_slot]
        # CHECKING IF THERE IS SUPPLY STUDENTS OR NOT
        try:
            ws_branch_sply = wb_branch[input_slot+'_supply']
        except:
            check_supply = 0
        # COUNT OF NORMAL STUDENTS
        for r in range(1, ws_branch_reg.max_row+1):
            noOfStudents += 1
        if (check_supply == 1):
            # COUNT OF SUPPLY STUDENTS
            for r in range(1, ws_branch_sply.max_row+1):
                noOfStudents += 1

# NO: OF STUDENTS WRITING EXAM FOR A PARTICULAR SLOT
print("No: of students for "+str(input_slot)+" :"+str(noOfStudents))
